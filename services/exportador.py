"""
Exportación de resultados a Excel y PDF.
PDF incluye logo, resumen general y una sección por ejecutivo.
"""
import os
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import CRITERIOS, FLAG_COLS, puntos_reciprocidad, RECIPROCIDAD_LABELS

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "asset", "critcom.png")
FIRMA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "asset", "firma_pablo.png")


def _nombre_flag(flag_key):
    return CRITERIOS.get(flag_key, {}).get("nombre", flag_key)


def _reemplazar_flags_en_texto(texto):
    if not isinstance(texto, str):
        return texto
    for fk, meta in CRITERIOS.items():
        texto = texto.replace(fk, meta["nombre"])
    return texto


def _preparar_df_export(df):
    cols_export = [
        "cuit", "nom_cliente", "tipo_empresa", "ubicacion_comercial",
        "estado", "total_flags_anterior", "total_flags_actual", "delta_flags",
        "cumplimiento_anterior", "cumplimiento_actual",
        "flags_ganados", "flags_perdidos",
        "criterio_nombre", "accion_tipo", "accion_texto", "score_nba",
        "cerca_nivel", "nombre_rol", "sucursal_rol", "tipo_rol",
    ]
    cols_disp = [c for c in cols_export if c in df.columns]
    df_exp = df[cols_disp].copy()

    for col in ["flags_ganados", "flags_perdidos"]:
        if col in df_exp.columns:
            df_exp[col] = df_exp[col].apply(_reemplazar_flags_en_texto)

    col_display = {
        "cuit": "CUIT", "nom_cliente": "Cliente", "tipo_empresa": "Tipo Empresa",
        "ubicacion_comercial": "Ubicación", "estado": "Estado",
        "total_flags_anterior": "Crit. Anterior", "total_flags_actual": "Crit. Actual",
        "delta_flags": "Delta", "cumplimiento_anterior": "Cumpl. Anterior",
        "cumplimiento_actual": "Cumpl. Actual",
        "flags_ganados": "Criterios que SUBIERON", "flags_perdidos": "Criterios que BAJARON",
        "criterio_nombre": "Criterio Sugerido", "accion_tipo": "Tipo Acción",
        "accion_texto": "Acción Sugerida", "score_nba": "Prioridad NBA",
        "cerca_nivel": "Quick Win Nivel",
        "nombre_rol": "Ejecutivo", "sucursal_rol": "Sucursal Rol",
        "tipo_rol": "Tipo Rol",
    }
    df_exp = df_exp.rename(columns=col_display)
    return df_exp


def exportar_excel(df, nombre_hoja="Comparación"):
    output = BytesIO()
    df_exp = _preparar_df_export(df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exp.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=1)
        ws = writer.sheets[nombre_hoja]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(df_exp.columns), 10))
        ws.cell(1, 1).value = f"CritCom - {nombre_hoja} - {datetime.now().strftime('%d/%m/%Y')}"
        ws.cell(1, 1).font = Font(bold=True, size=14, color="00A651")

        header_fill = PatternFill(start_color="00A651", end_color="00A651", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx in range(1, len(df_exp.columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        estado_col = None
        for col_idx, col_name in enumerate(df_exp.columns, 1):
            if col_name == "Estado":
                estado_col = col_idx
                break

        estado_colores = {
            "MEJORO": "C6EFCE", "EMPEORO": "FFC7CE", "NUEVO": "BDD7EE",
            "DESAPARECIDO": "F4B084", "SIN_CAMBIOS": "D9D9D9",
            "CAMBIO_LATERAL": "FFE699",
        }

        for row_idx in range(3, len(df_exp) + 3):
            for col_idx in range(1, len(df_exp.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)

            if estado_col:
                estado_val = ws.cell(row=row_idx, column=estado_col).value
                color = estado_colores.get(str(estado_val), "")
                if color:
                    for ci in range(1, len(df_exp.columns) + 1):
                        ws.cell(row=row_idx, column=ci).fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )

        for col_idx in range(1, len(df_exp.columns) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(ws.cell(2, col_idx).value or ""))
            for row_idx in range(3, min(len(df_exp) + 3, 50)):
                cell_len = len(str(ws.cell(row_idx, col_idx).value or ""))
                max_len = max(max_len, cell_len)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    output.seek(0)
    return output.getvalue()


def exportar_pdf(df, titulo="Reporte de Criterios Comerciales", indices_rol=None):
    """
    indices_rol: dict {nombre_rol: indice_float} para sección sugerencia liberar.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, Image, PageBreak,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return None

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    verde_bp = colors.HexColor("#00A651")
    cyan_bp = colors.HexColor("#00B8D4")
    gris_claro = colors.Color(0.96, 0.98, 0.96)

    s_titulo = ParagraphStyle("Titulo", parent=styles["Heading1"],
        fontSize=20, textColor=verde_bp, spaceAfter=4*mm, alignment=TA_CENTER)
    s_subtitulo = ParagraphStyle("Subtitulo", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey, spaceAfter=6*mm, alignment=TA_CENTER)
    s_seccion = ParagraphStyle("Seccion", parent=styles["Heading2"],
        fontSize=13, textColor=verde_bp, spaceBefore=5*mm, spaceAfter=3*mm)
    s_rol = ParagraphStyle("Rol", parent=styles["Heading2"],
        fontSize=14, textColor=colors.white, spaceAfter=2*mm, alignment=TA_CENTER)
    s_normal = ParagraphStyle("Normal2", parent=styles["Normal"], fontSize=8, leading=11)
    s_pie = ParagraphStyle("Pie", parent=styles["Normal"],
        fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    fecha_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ubicacion = ""
    if "ubicacion_comercial" in df.columns:
        ubs = df["ubicacion_comercial"].dropna().unique()
        ubicacion = ubs[0] if len(ubs) > 0 else ""

    def _calcular_puntos(df_sub):
        """Agrega columnas puntos_recip y nivel_recip al DataFrame."""
        df_sub = df_sub.copy()
        if "total_flags_actual" in df_sub.columns:
            df_sub["puntos_recip"] = df_sub["total_flags_actual"].apply(
                lambda x: puntos_reciprocidad(int(x or 0)))
            df_sub["nivel_recip"] = df_sub["puntos_recip"].map(RECIPROCIDAD_LABELS)
        else:
            df_sub["puntos_recip"] = 0
            df_sub["nivel_recip"] = "SIN"
        return df_sub

    def _criterios_texto(row, valor):
        """Genera texto con nombres de criterios que el cliente cumple (1) o no (0)."""
        nombres = []
        for f in FLAG_COLS:
            if f in row.index and f in CRITERIOS:
                if int(row.get(f, 0) or 0) == valor:
                    nombres.append(CRITERIOS[f]["nombre"])
        return ", ".join(nombres) if nombres else "-"

    def _tabla_clientes(df_sub, ordenar_por_puntos=False):
        """Genera flowable tabla de clientes para un segmento del PDF."""
        cols_pdf = ["nom_cliente", "cuit", "estado", "total_flags_actual",
                    "delta_flags", "cumple", "no_cumple", "criterio_nombre", "cerca_nivel"]
        # Agregar columnas calculadas
        df_sub = df_sub.copy()
        df_sub["cumple"] = df_sub.apply(lambda r: _criterios_texto(r, 1), axis=1)
        df_sub["no_cumple"] = df_sub.apply(lambda r: _criterios_texto(r, 0), axis=1)

        cols_use = [c for c in cols_pdf if c in df_sub.columns]
        hdrs = {
            "nom_cliente": "Cliente", "cuit": "CUIT", "estado": "Estado",
            "total_flags_actual": "Crit.", "delta_flags": "Delta",
            "cumple": "Cumple", "no_cumple": "No cumple",
            "criterio_nombre": "Sugerido", "cerca_nivel": "QW",
        }
        widths = {
            "nom_cliente": 4*cm, "cuit": 2.5*cm, "estado": 1.8*cm,
            "total_flags_actual": 1*cm, "delta_flags": 1*cm,
            "cumple": 5*cm, "no_cumple": 5*cm,
            "criterio_nombre": 3*cm, "cerca_nivel": 1.2*cm,
        }
        col_widths = [widths.get(c, 2*cm) for c in cols_use]
        data = [[hdrs.get(c, c) for c in cols_use]]
        color_map = {
            "MEJORO": colors.Color(0.78, 0.94, 0.81),
            "EMPEORO": colors.Color(1, 0.78, 0.81),
            "NUEVO": colors.Color(0.74, 0.84, 0.93),
            "DESAPARECIDO": colors.Color(0.96, 0.69, 0.51),
            "CAMBIO_LATERAL": colors.Color(1, 0.9, 0.6),
        }

        row_colors = []
        # Ordenar por puntos de reciprocidad (desc) o por score NBA
        if ordenar_por_puntos and "puntos_recip" in df_sub.columns:
            df_sorted = df_sub.sort_values("puntos_recip", ascending=False)
        elif "score_nba" in df_sub.columns:
            df_sorted = df_sub.sort_values("score_nba", ascending=False)
        else:
            df_sorted = df_sub
        for _, row in df_sorted[cols_use].iterrows():
            fila = []
            for c in cols_use:
                val = row[c]
                if pd.isna(val):
                    val = ""
                if c == "cerca_nivel":
                    val = "SI" if val else ""
                else:
                    val = str(val)
                    if len(val) > 45:
                        val = val[:42] + "..."
                fila.append(Paragraph(val, s_normal))
            data.append(fila)
            estado_val = str(row.get("estado", "")) if "estado" in cols_use else ""
            row_colors.append(color_map.get(estado_val, colors.white))

        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), verde_bp),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.Color(0.85, 0.85, 0.85)),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for i, bg in enumerate(row_colors, 1):
            estilo.append(("BACKGROUND", (0, i), (-1, i), bg))

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle(estilo))
        return t

    def _tabla_sugerencia_liberar(df_sub):
        """Tabla compacta de clientes que bajan el índice del rol."""
        cols = ["nom_cliente", "cuit", "total_flags_actual", "nivel_recip", "puntos_recip"]
        cols_use = [c for c in cols if c in df_sub.columns]
        hdrs = {"nom_cliente": "Cliente", "cuit": "CUIT", "total_flags_actual": "Criterios",
                "nivel_recip": "Reciprocidad", "puntos_recip": "Puntos"}
        widths_map = {"nom_cliente": 6*cm, "cuit": 3*cm, "total_flags_actual": 2*cm,
                      "nivel_recip": 3*cm, "puntos_recip": 2*cm}
        col_widths = [widths_map.get(c, 2.5*cm) for c in cols_use]

        data = [[hdrs.get(c, c) for c in cols_use]]
        df_sorted = df_sub.sort_values("puntos_recip", ascending=True)
        for _, row in df_sorted[cols_use].iterrows():
            fila = []
            for c in cols_use:
                val = row[c]
                if pd.isna(val):
                    val = ""
                fila.append(Paragraph(str(val), s_normal))
            data.append(fila)

        naranja = colors.Color(0.96, 0.85, 0.65)
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E67E22")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.Color(0.85, 0.85, 0.85)),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for i in range(1, len(data)):
            estilo.append(("BACKGROUND", (0, i), (-1, i), naranja))

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle(estilo))
        return t

    def _tabla_desaparecidos(df_sub):
        """Tabla de clientes que dejaron de aparecer en el listado."""
        cols = ["nom_cliente", "cuit", "total_flags_anterior", "cumplimiento_anterior"]
        cols_use = [c for c in cols if c in df_sub.columns]
        hdrs = {"nom_cliente": "Cliente", "cuit": "CUIT",
                "total_flags_anterior": "Crit. Anterior", "cumplimiento_anterior": "Cumpl. Anterior"}
        widths_map = {"nom_cliente": 6*cm, "cuit": 3*cm,
                      "total_flags_anterior": 3*cm, "cumplimiento_anterior": 4*cm}
        col_widths = [widths_map.get(c, 2.5*cm) for c in cols_use]

        data = [[hdrs.get(c, c) for c in cols_use]]
        for _, row in df_sub[cols_use].iterrows():
            fila = [Paragraph(str(row[c]) if not pd.isna(row[c]) else "", s_normal) for c in cols_use]
            data.append(fila)

        gris_naranja = colors.Color(0.96, 0.69, 0.51)
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F4B084")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.Color(0.85, 0.85, 0.85)),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle(estilo))
        return t

    def _header_rol(nombre_rol, n_clientes, elems):
        """Agrega un banner verde con el nombre del ejecutivo."""
        data_banner = [[Paragraph(f"{nombre_rol}  —  {n_clientes} clientes", s_rol)]]
        t = Table(data_banner, colWidths=[25*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), verde_bp),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elems.append(t)
        elems.append(Spacer(1, 3*mm))

    elementos = []

    # ── Portada / Resumen general ─────────────────────────────────────────
    # Logo
    logo_path = os.path.normpath(LOGO_PATH)
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=4*cm, height=4*cm, kind="proportional")
            img.hAlign = "CENTER"
            elementos.append(img)
            elementos.append(Spacer(1, 3*mm))
        except Exception:
            pass

    elementos.append(Paragraph(titulo, s_titulo))
    elementos.append(Paragraph(
        f"Banco Provincia | {ubicacion} | Generado: {fecha_str}",
        s_subtitulo,
    ))
    elementos.append(HRFlowable(width="100%", thickness=2, color=verde_bp, spaceAfter=5*mm))

    # KPIs
    estados = df["estado"].value_counts() if "estado" in df.columns else pd.Series(dtype=int)
    cerca_nivel_n = int(df["cerca_nivel"].sum()) if "cerca_nivel" in df.columns else 0
    kpi_data = [
        ["Total", "Nuevos", "Mejoraron", "Empeoraron", "Sin Cambios", "Desaparecidos", "Quick Win Nivel"],
        [
            str(len(df)),
            str(estados.get("NUEVO", 0)),
            str(estados.get("MEJORO", 0)),
            str(estados.get("EMPEORO", 0)),
            str(estados.get("SIN_CAMBIOS", 0)),
            str(estados.get("DESAPARECIDO", 0)),
            str(cerca_nivel_n),
        ],
    ]
    t_kpi = Table(kpi_data, colWidths=[3.6*cm]*7)
    t_kpi.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), verde_bp),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (6, 0), (6, -1), cyan_bp),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, 1), 16),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(t_kpi)
    elementos.append(Spacer(1, 6*mm))

    # Top oportunidades generales
    elementos.append(Paragraph("Top oportunidades generales (por score NBA)", s_seccion))
    df_top = df.copy()
    if "score_nba" in df_top.columns:
        df_top = df_top[df_top["score_nba"] > 0].nlargest(20, "score_nba")
    if not df_top.empty:
        elementos.append(_tabla_clientes(df_top))

    # Pie de página del resumen
    elementos.append(Spacer(1, 8*mm))
    elementos.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    elementos.append(Paragraph(
        f"CritCom — Banco Provincia | Generado: {fecha_str}", s_pie
    ))

    # ── Sección por ejecutivo ─────────────────────────────────────────────
    if "nombre_rol" in df.columns:
        roles = [r for r in df["nombre_rol"].dropna().unique() if r and r != "SIN ASIGNAR"]
        roles_sorted = sorted(roles, key=lambda r: df[df["nombre_rol"] == r]["cuit"].count(), reverse=True)

        # Agregar puntos de reciprocidad a todo el DataFrame
        df = _calcular_puntos(df)

        for rol in roles_sorted:
            df_rol = df[df["nombre_rol"] == rol].copy()
            if df_rol.empty:
                continue

            elementos.append(PageBreak())

            # Banner con nombre del rol
            _header_rol(rol, len(df_rol), elementos)

            # KPIs del ejecutivo
            est_rol = df_rol["estado"].value_counts()
            cerca_rol = int(df_rol["cerca_nivel"].sum()) if "cerca_nivel" in df_rol.columns else 0
            prom_crit = df_rol["total_flags_actual"].mean() if "total_flags_actual" in df_rol.columns else 0

            kpi_rol = [
                ["Clientes", "Mejoraron", "Empeoraron", "Nuevos", "Prom. Criterios", "Quick Win Nivel"],
                [
                    str(len(df_rol)),
                    str(est_rol.get("MEJORO", 0)),
                    str(est_rol.get("EMPEORO", 0)),
                    str(est_rol.get("NUEVO", 0)),
                    f"{prom_crit:.1f} / 15",
                    str(cerca_rol),
                ],
            ]
            t_rol = Table(kpi_rol, colWidths=[3.6*cm, 3.6*cm, 3.6*cm, 3.6*cm, 4*cm, 4*cm])
            t_rol.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), verde_bp),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (5, 0), (5, -1), cyan_bp),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, 1), 14),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elementos.append(t_rol)
            elementos.append(Spacer(1, 4*mm))

            # Tabla de clientes del ejecutivo (ordenada por puntos desc)
            df_rol_activos = df_rol[df_rol["estado"] != "DESAPARECIDO"]
            if not df_rol_activos.empty:
                elementos.append(_tabla_clientes(df_rol_activos, ordenar_por_puntos=True))

            # Sección: clientes por debajo del índice
            if "puntos_recip" in df_rol.columns and indices_rol:
                indice_rol = indices_rol.get(rol)
                if indice_rol is not None:
                    df_bajan = df_rol[
                        (df_rol["puntos_recip"] < indice_rol) &
                        (df_rol["estado"] != "DESAPARECIDO")
                    ]
                    if not df_bajan.empty:
                        elementos.append(Spacer(1, 4*mm))
                        elementos.append(Paragraph(
                            "Clientes por debajo de tu índice — revisá si conviene mejorarlos o liberarlos",
                            s_seccion))
                        elementos.append(_tabla_sugerencia_liberar(df_bajan))

            # Sección: desaparecidos del ejecutivo
            df_desap_rol = df_rol[df_rol["estado"] == "DESAPARECIDO"]
            if not df_desap_rol.empty:
                elementos.append(Spacer(1, 4*mm))
                elementos.append(Paragraph(
                    f"Clientes que dejaron de aparecer en el listado ({len(df_desap_rol)})",
                    s_seccion))
                elementos.append(_tabla_desaparecidos(df_desap_rol))

            # Pie
            elementos.append(Spacer(1, 5*mm))
            elementos.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
            elementos.append(Paragraph(
                f"CritCom — {rol} | {fecha_str}", s_pie
            ))

        # Sección sin asignar (ordenada por total_flags desc)
        df_sin = df[df["nombre_rol"].fillna("SIN ASIGNAR") == "SIN ASIGNAR"].copy()
        if not df_sin.empty:
            if "total_flags_actual" in df_sin.columns:
                df_sin = df_sin.sort_values("total_flags_actual", ascending=False)
            elementos.append(PageBreak())
            _header_rol("SIN ASIGNAR", len(df_sin), elementos)
            elementos.append(_tabla_clientes(df_sin, ordenar_por_puntos=True))

            # Desaparecidos sin asignar
            df_desap_sin = df_sin[df_sin["estado"] == "DESAPARECIDO"]
            if not df_desap_sin.empty:
                elementos.append(Spacer(1, 4*mm))
                elementos.append(Paragraph(
                    f"Sin asignar — dejaron de aparecer ({len(df_desap_sin)})",
                    s_seccion))
                elementos.append(_tabla_desaparecidos(df_desap_sin))

            elementos.append(Spacer(1, 5*mm))
            elementos.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
            elementos.append(Paragraph(f"CritCom — SIN ASIGNAR | {fecha_str}", s_pie))

    doc.build(elementos)
    output.seek(0)
    return output.getvalue()
